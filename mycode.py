import os
import json
import csv, io, requests
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import openpyxl
import fitz  # PyMuPDF

CONFIG_FILE = "config.txt"

# ----------------- HỖ TRỢ -----------------
def normalize_rgb(values):
    """Chuyển đổi hex (#RRGGBB) hoặc (R,G,B) sang tuple (0-1)"""
    if isinstance(values[0], str) and values[0].startswith("#"):
        hexcode = values[0].lstrip("#")
        if len(hexcode) == 6:
            r = int(hexcode[0:2], 16) / 255
            g = int(hexcode[2:4], 16) / 255
            b = int(hexcode[4:6], 16) / 255
            return (r, g, b)
    nums = []
    for v in values:
        if v is None:
            continue
        v = float(v)
        nums.append(v/255 if v > 1 else v)
    if len(nums) == 3:
        return tuple(nums)
    raise ValueError(f"Không hợp lệ: {values}")

def rgb_to_hex(rgb):
    return "#{:02X}{:02X}{:02X}".format(
        int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255)
    )

# ----------------- ĐỌC MAPPING -----------------
def get_mapping_from_excel(path):
    mapping = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheetname in ["Cable", "Device"]:
        if sheetname not in wb.sheetnames:
            continue
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            subject = str(row[0]).strip().upper()
            hexcode = str(row[1]).strip()
            note = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            try:
                rgb = normalize_rgb([hexcode])
                mapping[subject] = (rgb, note)
            except Exception as e:
                print(f"⚠️ Bỏ qua {row}: {e}")
                continue
    return mapping

def get_mapping_from_gsheet(url, note_col=2):
    """Đọc Google Sheet CSV export"""
    if "/edit" in url and "gid=" in url:
        parts = url.split("gid=")
        gid = parts[1].split("&")[0] if "&" in parts[1] else parts[1]
        doc_id = url.split("/d/")[1].split("/")[0]
        url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv&gid={gid}"

    mapping = {}
    try:
        response = requests.get(url)
        response.raise_for_status()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không tải được Google Sheet:\n{e}")
        return mapping

    reader = csv.reader(io.StringIO(response.text))
    next(reader, None)

    for row in reader:
        if not row or not row[0]:
            continue
        subject = row[0].strip().upper()
        try:
            if row[1].startswith("#"):
                rgb = normalize_rgb([row[1]])
            else:
                rgb = normalize_rgb(row[1:4])
            note = row[note_col].strip() if len(row) > note_col else ""
            mapping[subject] = (rgb, note)
        except Exception as e:
            print(f"⚠️ Bỏ qua {row}: {e}")
            continue
    return mapping

# ----------------- QUẢN LÝ CONFIG -----------------
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                content = f.read().strip()
                if content:
                    return json.loads(content)
                else:
                    print("⚠️ config.txt rỗng.")
        except Exception as e:
            print(f"⚠️ Lỗi đọc config.txt: {e}")
    return {}

def save_config(config):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

# ----------------- MAIN -----------------
def main():
    root = tk.Tk()
    root.withdraw()

    config = load_config()

    # B1: PDF
    pdf_paths = config.get("pdf_paths", [])
    if not pdf_paths:
        pdf_paths = filedialog.askopenfilenames(
            title="Chọn một hoặc nhiều file PDF",
            filetypes=[("PDF Files", "*.pdf")]
        )
        pdf_paths = list(pdf_paths)
        config["pdf_paths"] = pdf_paths
        save_config(config)
    if not pdf_paths:
        messagebox.showinfo("Thông báo", "Không có file PDF nào được chọn.")
        return

    # B2: Mapping
    mapping = {}
    if config.get("type") == "gsheet":
        if config.get("url_cable"):
            mapping.update(get_mapping_from_gsheet(config["url_cable"]))
        if config.get("url_device"):
            mapping.update(get_mapping_from_gsheet(config["url_device"]))
    elif config.get("type") == "excel":
        if config.get("excel_path"):
            mapping = get_mapping_from_excel(config["excel_path"])

    if not mapping:
        # Nếu chưa có mapping thì hỏi
        choice = messagebox.askyesno(
            "Chọn nguồn mapping",
            "Bạn có muốn dùng Google Sheet không?\nYes = Google Sheet, No = Excel"
        )

        if choice:
            url_cable = simpledialog.askstring("Google Sheet - Cable", "Nhập link Google Sheet cho Cable:")
            url_device = simpledialog.askstring("Google Sheet - Device", "Nhập link Google Sheet cho Device:")
            config["type"] = "gsheet"
            config["url_cable"] = url_cable
            config["url_device"] = url_device
            if url_cable:
                mapping.update(get_mapping_from_gsheet(url_cable))
            if url_device:
                mapping.update(get_mapping_from_gsheet(url_device))
        else:
            excel_path = filedialog.askopenfilename(
                title="Chọn file Excel (Cable_Device_Table.xlsx)",
                filetypes=[("Excel files", "*.xlsx *.xlsm")]
            )
            config["type"] = "excel"
            config["excel_path"] = excel_path
            mapping = get_mapping_from_excel(excel_path)

        save_config(config)

    if not mapping:
        messagebox.showerror("Lỗi", "Không có mapping hợp lệ!")
        return

    print("\n===== KẾT QUẢ MAPPING =====")
    for k, (v, note) in mapping.items():
        print(f"{k:20} → {v} ({rgb_to_hex(v)}) | Note: {note}")
    print("============================\n")

    # B3: Xử lý PDF
    for pdf_path in pdf_paths:
        doc = fitz.open(pdf_path)
        for page_num, page in enumerate(doc, start=1):
            annots = page.annots()
            if not annots:
                continue

            for annot in annots:
                subj_candidates = []
                for key in ["title", "subject", "content"]:
                    try:
                        v = annot.info.get(key, "")
                        if v:
                            subj_candidates.append(v)
                    except Exception:
                        continue

                for subj in subj_candidates:
                    subj_norm = subj.strip().upper()
                    if subj_norm in mapping:
                        color, note = mapping[subj_norm]
                        print(f"🔄 {os.path.basename(pdf_path)} - Trang {page_num}: "
                              f"Đổi {subj_norm} sang {rgb_to_hex(color)} | Note: {note}")

                        annot.set_colors(stroke=color, fill=color)
                        annot.update()

                        if note:
                            rect = annot.rect
                            char_width = 5
                            box_width = max(60, len(note) * char_width)
                            box_height = 15

                            note_upper = note.upper()
                            if "GLANDED (FROM)" in note_upper or "TERMINATED (FROM)" in note_upper:
                                note_rect = fitz.Rect(
                                    rect.x0 - box_width - 5,
                                    rect.y0,
                                    rect.x0 - 5,
                                    rect.y0 + box_height
                                )
                            else:
                                note_rect = fitz.Rect(
                                    rect.x1 + 5,
                                    rect.y0,
                                    rect.x1 + 5 + box_width,
                                    rect.y0 + box_height
                                )

                            page.add_freetext_annot(
                                note_rect,
                                note,
                                fontsize=9,
                                fontname="helv",
                                rotate=0,
                                text_color=color
                            )
                        break

        output_path = pdf_path.replace(".pdf", "_colored.pdf")
        doc.save(output_path)
        doc.close()
        print(f"✅ Đã lưu file mới: {output_path}")

    messagebox.showinfo("Hoàn tất", "Đã xử lý xong tất cả PDF!")

# ----------------- RUN -----------------
if __name__ == "__main__":
    main()
